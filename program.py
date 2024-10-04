import os
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# Load credentials 
username = 'rajkumarsain.doit'
password = 'Network@1984'

# Get current directory and ensure Excel file is in the same directory as Python script
current_dir = os.getcwd()
excel_filename = input("Please enter the Excel filename (e.g., 'data.xlsx'): ").strip()
excel_path = os.path.join(current_dir, excel_filename)

# Check if the file exists
if not os.path.isfile(excel_path):
    print(f"Error: The file '{excel_filename}' does not exist in the current directory.")
    exit(1)

# Set up Chrome WebDriver using Service
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

# Step 1: Open the login page
driver.get("https://sso.rajasthan.gov.in/signin")

# Step 2: Login to the portal
WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.ID, 'cpBody_cpBody_txt_Data1'))
).send_keys(username)

WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.ID, 'cpBody_cpBody_txt_Data2'))
).send_keys(password)

WebDriverWait(driver, 20).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="cpBody_cpBody_btn_LDAPLogin"]'))
).click()

# Step 3: Wait for login to complete and navigate
WebDriverWait(driver, 20).until(
    EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'h3.filterable'))
)

try:
    # Locate and click "JAN AADHAAR"
    h3_elements = WebDriverWait(driver, 20).until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'h3.filterable'))
    )
    for element in h3_elements:
        if "JAN AADHAAR" in element.text.upper():
            element.click()
            break
    else:
        print("JAN AADHAAR not found")
        exit(1)

    # Wait for page load and click ENROLLMENT and GENERIC SEARCH
    WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, 'span.dashOptEnrolementEdit[title="#"]'))
    ).click()

    WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="div_menuId_1202"]//span[@class="dashOptPopup"]'))
    ).click()

    # Load Excel data and process in batches
    df = pd.read_excel(excel_path)  # Load the Excel file into a DataFrame
    wb = load_workbook(excel_path)  # Open the workbook for updating
    ws = wb.active

    batch_size = 100  # Process 100 records at a time
    total_records = len(df)

    for batch_start in range(0, total_records, batch_size):
        batch_end = min(batch_start + batch_size, total_records)
        batch_df = df.iloc[batch_start:batch_end]

        for row_index, (index, row) in enumerate(batch_df.iterrows(), start=batch_start + 2):
            # Check if the AADHAR column is empty
            aadhar_number = str(row['AADHAR']).strip() if pd.notna(row['AADHAR']) else None
             # Only proceed if AADHAR is empty
            if not aadhar_number:
                # Read the mobile number only if AADHAR is empty
                mobile_number = str(row['MOBILE_NO']).strip() if pd.notna(row['MOBILE_NO']) else None
                name_from_excel = row['NAME'].strip() if pd.notna(row['NAME']) else None
                
                # Use row_index for Excel row calculation directly
                excel_row = row_index  # Adjust for header row (starts from 2)

                # Check for missing data
                if not mobile_number or not name_from_excel:
                    ws.cell(row=excel_row, column=df.columns.get_loc('AADHAR') + 1).value = "Missing data"
                    continue

                # Step 7: Enter mobile number and search
                mobile_input = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.ID, 'mobileNo'))
                )
                mobile_input.clear()
                mobile_input.send_keys(mobile_number)

                search_button = driver.find_element(By.ID, 'btn')
                search_button.click()

                # Wait for the search results
                try:
                    WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.XPATH, "//*[@id='showdata']/table"))
                    )
                except Exception:
                    error_message = f"No Records Found {mobile_number}, skipping."
                    print(error_message)
                    # Write the error message in the AADHAR column of the corresponding row
                    ws.cell(row=excel_row, column=df.columns.get_loc('AADHAR') + 1).value = "No records Found"
                    continue
                # Check for "No Records Found"
                try:
                    not_found_message = driver.find_element(By.XPATH, "//div[contains(text(), 'No Records Found')]")
                    if not_found_message:
                        ws.cell(row=excel_row, column=df.columns.get_loc('AADHAR') + 1).value = "No Records Found"
                        print(f"Mobile number {mobile_number}: No records found, skipping.")
                        continue
                except Exception:
                    pass  # Continue if "No Records Found" is not present

                # Match name from webpage
                try:
                    rows_xpath = "//*[@id='showdata']/table/tbody/tr"
                    rows = WebDriverWait(driver, 20).until(
                        EC.presence_of_all_elements_located((By.XPATH, rows_xpath))
                    )

                    match_found = False
                    for table_row in rows:
                        columns = table_row.find_elements(By.TAG_NAME, 'td')

                        if len(columns) >= 4:
                            name_on_web = columns[3].text.strip()
                            aadhar_id = columns[2].text.strip() if columns[2].text.strip() else "aadhar not available"

                            if name_from_excel.lower() == name_on_web.lower():
                                ws.cell(row=excel_row, column=df.columns.get_loc('AADHAR') + 1).value = aadhar_id
                                ws.cell(row=excel_row, column=df.columns.get_loc('WebName') + 1).value = name_on_web
                                match_found = True
                                break

                    if not match_found:
                        ws.cell(row=excel_row, column=df.columns.get_loc('AADHAR') + 1).value = "No match found"

                except Exception as e:
                    ws.cell(row=excel_row, column=df.columns.get_loc('AADHAR') + 1).value = "Error during match"
                    print(f"Error processing record {index + 1}: {e}")

        # Save the Excel file after each batch
        wb.save(excel_path)
        print(f"Processed batch {batch_start + 1} to {batch_end}")

except Exception as e:
    print(f"Error: {e}")

# Final save and close workbook
wb.save(excel_path)
wb.close()

print("Processing complete.")
input("Press Enter to exit and close the browser...")
driver.quit()