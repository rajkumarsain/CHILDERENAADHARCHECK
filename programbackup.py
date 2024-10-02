import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# Load credentials 
username = 'rajkumarsain.doit'
password = 'Network@1984'

# Take Excel file path as input
excel_path = input("Please enter the path of the Excel file: ")

# Set up Chrome WebDriver using Service
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

# Step 1: Open the login page
driver.get("https://sso.rajasthan.gov.in/signin")

# Step 2: Login to the portal
time.sleep(2)  # Adjust sleep time based on the page loading time
username_field = driver.find_element(By.ID, 'cpBody_cpBody_txt_Data1')
username_field.send_keys(username)
password_field = driver.find_element(By.ID, 'cpBody_cpBody_txt_Data2')
password_field.send_keys(password)
login_button = driver.find_element(By.XPATH, '//*[@id="cpBody_cpBody_btn_LDAPLogin"]')
time.sleep(10)
login_button.click()

# Step 3: Wait for login to complete and navigate
time.sleep(10)

try:
    # Locate and click "JAN AADHAAR"
    h3_elements = WebDriverWait(driver, 20).until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'h3.filterable'))
    )
    for element in h3_elements:
        if "JAN AADHAAR" in element.text.upper():
            element.click()
            break
    else:
        print("JAN AADHAAR not found")
    
    time.sleep(5)
    
    # Step 6: Click on ENROLLMENT and GENERIC SEARCH
    WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, 'span.dashOptEnrolementEdit[title="#"]'))
    ).click()
    WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="div_menuId_1202"]//span[@class="dashOptPopup"]'))
    ).click()

    # Step 8: Process Excel data in batches
    df = pd.read_excel(excel_path)
    wb = load_workbook(excel_path)
    ws = wb.active

    batch_size = 50  # Define batch size to process records in smaller chunks
    total_records = len(df)
    for batch_start in range(0, total_records, batch_size):
        batch_end = min(batch_start + batch_size, total_records)
        batch_df = df.iloc[batch_start:batch_end]

        for index, row in batch_df.iterrows():
            if pd.isna(row['MOBILE_NO']):
                continue  # Skip blank mobile numbers
            else:
                mobile_number = str(row['MOBILE_NO']).strip()
                name_from_excel = row['NAME'].strip()
                
                # Step 7: Enter mobile number and search
                mobile_input = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.ID, 'mobileNo'))
                )
                mobile_input.clear()
                mobile_input.send_keys(mobile_number)
                search_button = driver.find_element(By.ID, 'btn')
                search_button.click()
                time.sleep(5)  # Wait for the search results
                try:
                    # Match name from webpage
                    table_xpath = "//*[@id='showdata']/table"
                    WebDriverWait(driver, 20).until(
                        EC.visibility_of_element_located((By.XPATH, table_xpath))
                    )

                    rows_xpath = "//*[@id='showdata']/table/tbody/tr"
                    rows = WebDriverWait(driver, 20).until(
                        EC.presence_of_all_elements_located((By.XPATH, rows_xpath))
                    )

                    match_found = False
                    for row_index, row in enumerate(rows, start=1):
                        columns = row.find_elements(By.TAG_NAME, 'td')

                        if len(columns) >= 4:
                            name_on_web = columns[3].text.strip()
                            aadhar_id = columns[2].text.strip()

                            if name_from_excel.lower() == name_on_web.lower():
                                ws.cell(row=batch_start + index + 2, column=df.columns.get_loc('AADHAR') + 1).value = aadhar_id
                                match_found = True
                                break

                    if not match_found:
                        ws.cell(row=batch_start + index + 2, column=df.columns.get_loc('AADHAR') + 1).value = "No match found"
                except Exception as e:
                  
                    ws.cell(row=batch_start + index + 2, column=df.columns.get_loc('AADHAR') + 1).value = "Error during match"
    

        # Save the Excel file after processing each batch
        wb.save(excel_path)
        print(f"Processed batch {batch_start + 1} to {batch_end}")

except Exception as e:
    print(f"Error: {e}")

# Final save and close workbook
wb.save(excel_path)
wb.close()

input("Press Enter to exit and close the browser...")
# Optionally close the browser
driver.quit()
